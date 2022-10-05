# Import Semua Modules
import pandas as pd                                      
from openpyxl import load_workbook
from openpyxl.styles import *
from openpyxl.chart import *
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList

input_file      = 'data_input/supermarket_sales.xlsx'
output_file     = 'data_output/report_penjualan_2019.xlsx'
webhook_url     = 'https://discord.com/api/webhooks/1027166345856552991/Ybh4K8urJ7iqKG__pQWK-ODvXr2-Ak8vPXDmXkGWzNtgvKvrh7hYd8pHu1aTfizvfQva'


# Baca Data, buat pivot table, dan simpan file baru
df = pd.read_excel(input_file)

df = df.pivot_table(index='Gender', 
                    columns='Product line', 
                    values='Total', 
                    aggfunc='sum').round()

print('Save dataframe to excel...')

df.to_excel(output_file, 
                sheet_name='Report', 
                startrow=4)

print('Save dataframe done...')

# Membuat Grafik dari pivot table
wb = load_workbook(output_file)
wb.active = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#print(min_column, max_column, min_row, max_row)

barchart = BarChart()

data = Reference(wb.active, 
                min_col=min_column+1,
                max_col=max_column,
                min_row=min_row,
                max_row=max_row
                )

categories = Reference(wb.active,
                        min_col=min_column,
                        max_col=max_column,
                        min_row=min_row+1,
                        max_row=max_row
                        )

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

wb.active.add_chart(barchart, 'B12')
barchart.title = 'Sales berdasarkan Produk'
barchart.style = 2
wb.save(output_file)

# Menambah total penjualan dan memberi kop sederhana pada laporan
import string
alphabet = list(string.ascii_uppercase)
alphabet_excel = alphabet[:max_column]
#[A,B,C,D,E,F,G]
for i in alphabet_excel:
    if i != 'A':
        wb.active[f'{i}{max_row+1}'] = f'=SUM({i}{min_row+1}:{i}{max_row})'
        wb.active[f'{i}{max_row+1}'].style = 'Currency'

wb.active[f'{alphabet_excel[0]}{max_row+1}'] = 'Total'

wb.active['A1'] = 'Sales Report'
wb.active['A2'] = '2019'
wb.active['A1'].font = Font('Arial', bold=True, size=20)
wb.active['A2'].font = Font('Arial', bold=True, size=10)

wb.save(output_file)

# Mengirim ke Discord yang sudah di define
def send_to_discord():
    import discord
    from discord import SyncWebhook

    webhook = SyncWebhook.from_url(webhook_url)

    with open(file=output_file, mode='rb') as file:
        excel_file = discord.File(file)

    webhook.send('This is an automated report', 
                username='Report-Penjualan', 
                file=excel_file)

send_to_discord()